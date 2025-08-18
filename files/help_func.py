import os
import platform
import shutil
import subprocess
import time
import tempfile
import psutil
import openpyxl
import threading
import stat
from pathlib import Path
import firebird.driver as fdb
from firebird.driver import connect_server, SrvInfoCode   


def get_pom_file():
    if platform.system() == "Windows":
        TEST_DIR = os.environ.get('TEST_DIR')
        VERSION = os.environ.get('VERSION')
        if VERSION and TEST_DIR:
            file = os.path.join(TEST_DIR, f"RedExpert-{VERSION}/pom.xml")
            with open(file, 'r') as f:
                context = f.read()
                print(context)

def kill_redexpert():
    time.sleep(10)
    for proc in psutil.process_iter():
        if proc.name() == f'RedExpert64{get_exe()}' or proc.name() == f'java{get_exe()}':
            proc.terminate()

def run_server():
    global p 
    p = subprocess.Popen([os.environ.get('PYTHON', 'python'), "./files/run_server.py"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    time.sleep(5)

def run_isql():
    def target():
        global p
        home_directory, _, _ = get_server_info()
        p = subprocess.Popen([f"{home_directory}/isql{get_exe()}","localhost:employee.fdb", "-q", "-u", "TEST_USER", "-p", "123", "-r", "TEST_ROLE"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)        
        p.wait()

    thread = threading.Thread(target=target)  
    thread.daemon = True
    thread.start()

def stop_server():
    global p
    p.terminate()
    p = None

def get_build_no():
    return os.environ.get('BUILD', "202506")

def backup_savedconnections_file():
    backup_file("connection-saved.xml")

def restore_savedconnections_file():
    restore_file("connection-saved.xml")

def backup_user_properties():
    backup_file("user.properties")

def restore_user_properties():
    restore_file("user.properties")

def backup_audit_profiles():
    backup_file("audit.profiles.config")

def restore_audit_profiles():
    restore_file("audit.profiles.config")

def backup_file(file_name: str):
    home_dir = os.path.expanduser("~")
    build_no = get_build_no()
    file_path = os.path.join(home_dir, f'.redexpert/{build_no}/{file_name}')
    shutil.copy(file_path, file_path + ".bak")

def restore_file(file_name: str):
    home_dir = os.path.expanduser("~")
    build_no = get_build_no()
    file_path = os.path.join(home_dir, f'.redexpert/{build_no}/{file_name}')
    if os.path.exists(file_path + ".bak"):
        if os.path.exists(file_path):
            os.remove(file_path)
        shutil.move(file_path + ".bak", file_path)

def set_urls(urls: str):
    home_dir = os.path.expanduser("~")
    build_no = get_build_no()
    user_properties_file = os.path.join(home_dir, f'.redexpert/{build_no}/user.properties')
    with open(user_properties_file, 'r') as f:
        context = f.read()

    context += urls
    
    with open(user_properties_file, 'w') as f:
        f.write(context)

def get_path_to_lib():
    return os.environ.get('DIST', 'C:\\Program Files\\RedExpert') + "/lib"

def get_path():
    DIST = os.environ.get('DIST', 'C:\\Program Files\\RedExpert')
    COVERAGE = os.environ.get('COVERAGE')
    bin = get_exe()
    if COVERAGE:
        path_to_exe = f"java -javaagent:./lib/jacocoagent.jar=destfile=./results/jacoco.exec,output=file -jar {DIST}/RedExpert.jar -exe_path={DIST}/bin/RedExpert64{bin}"
    else:
        path_to_exe = f"{DIST}/bin/RedExpert64{bin}"
    return path_to_exe

def clear_history_files():
    home_dir = os.path.expanduser("~")
    build_no = get_build_no()
    history_file = os.path.join(home_dir, f'.redexpert/{build_no}/connection-history.xml')
    shortcuts_file = os.path.join(home_dir, f'.redexpert/{build_no}/shortcuts.properties')
    user_panel_state_file = os.path.join(home_dir, f'.redexpert/{build_no}/saved-values.xml')
    query_dir = os.path.join(home_dir, f'.redexpert/editor')
    if os.path.exists(query_dir):
        shutil.rmtree(query_dir)
    for file in [history_file, shortcuts_file, user_panel_state_file]:
        if os.path.exists(file):
            os.remove(file)

def get_hosts_history_file():
    home_dir = os.path.expanduser("~")
    hosts_history_file = os.path.join(home_dir, '.redexpert/hosts.history')
    return hosts_history_file

def copy_dist_path():
    def _on_rm_error(func, path, exc_info):
        os.chmod(path, stat.S_IWRITE)
        func(path)

    DIST = os.environ.get('DIST', "C:\\Program Files\\RedExpert")
    tmp_dir = tempfile.gettempdir()

    if os.path.exists(tmp_dir + '/RedExpert'):
        shutil.rmtree(tmp_dir + '/RedExpert', onerror=_on_rm_error)
    return_path = shutil.copytree(DIST, tmp_dir + '/RedExpert')
    path_to_exe = return_path + f"/bin/RedExpert64{get_exe()}"
    return path_to_exe

def get_server_info():    
    if is_rdb26():
        home_directory = "/opt/RedDatabase/" if platform.system() == "Linux" else "C:\\RedDatabase(x64)\\"
        version = "2.6"
        srv_version = "RedDatabase"
    else:        
        with connect_server(server='localhost', user='SYSDBA', password='masterkey') as srv:
            home_directory = srv.info.home_directory
            version = str(srv._engine_version())
            srv_version = next(ver for ver in ["Firebird", "RedDatabase"] if srv.info.get_info(SrvInfoCode.SERVER_VERSION).find(ver) > -1)
    
    return home_directory, version, srv_version

def lock_employee():
    home_directory, version, srv_version = get_server_info()
    bin_dir = "bin/" if platform.system() == "Linux" or is_rdb26() else ""
    subprocess.run([f"{home_directory}{bin_dir}nbackup{get_exe()}", "-L", f"{home_directory}examples/empbuild/employee.fdb", "-u", "SYSDBA", "-p", "masterkey"])
    time.sleep(1)

def unlock_employee():
    home_directory, version, srv_version = get_server_info()
    bin_dir = "bin/" if platform.system() == "Linux" or is_rdb26() else ""
    delta_file = home_directory + "examples/empbuild/employee.fdb.delta"
    if os.path.exists(delta_file):
        time.sleep(2) 
        os.remove(delta_file)
        subprocess.run([f"{home_directory}{bin_dir}nbackup{get_exe()}", "-F", f"{home_directory}examples/empbuild/employee.fdb"])

def execute(query: str):
    """
    Execute select query and return fetchall result
    """
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
    with fdb.connect("localhost:employee.fdb", user='SYSDBA', password='masterkey') as con:
        cur = con.cursor()
        cur.execute(query)
        result = cur.fetchall()
        cur.close()
    return str(result)

def execute_immediate(query: str):
    """
    Execute query without returning result
    """
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
    with fdb.connect("localhost:employee.fdb", user='SYSDBA', password='masterkey') as con:
        con.execute_immediate(query)
        con.commit()

def delete_query_files():
    home_dir = os.path.expanduser("~")
    for path in Path(os.path.join(home_dir, f'.redexpert/editor')).glob("script*.sql"):
        os.remove(path)

def check_build_config(conf_path: str, number: int):
    check_dict = {}   
    check_dict["format"] = ["text", "text"]
    check_dict["enabled"] = ["true", "true"]
    check_dict["log_security_incidents"] = ["true", "false"]
    check_dict["log_initfini"] = ["true", "false"]
    check_dict["log_connections"] = ["true", "false"]
    check_dict["log_transactions"] = ["true", "false"]
    check_dict["log_statement_prepare"] = ["true", "false"]
    check_dict["log_statement_free"] = ["true", "false"]
    check_dict["log_statement_start"] = ["true", "false"]
    check_dict["log_statement_finish"] = ["true", "false"]
    check_dict["log_procedure_start"] = ["true", "false"]
    check_dict["log_procedure_finish"] = ["true", "false"]
    check_dict["log_procedure_compile"] = ["false", "true"]
    check_dict["log_function_start"] = ["true", "false"]
    check_dict["log_function_finish"] = ["true", "false"]
    check_dict["log_function_compile"] = ["true", "true"]
    check_dict["log_trigger_start"] = ["true", "false"]
    check_dict["log_service_query"] = ["false", "false"]
    check_dict["log_trigger_finish"] = ["false", "true"]
    check_dict["log_trigger_compile"] = ["true", "false"]
    check_dict["log_context"] = ["false", "true"]
    check_dict["log_errors"] = ["false", "true"]
    check_dict["log_warnings"] = ["false", "true"]
    check_dict["print_plan"] = ["false", "true"]
    check_dict["print_perf"] = ["false", "true"]
    check_dict["log_blr_requests"] = ["false", "true"]
    check_dict["print_blr"] = ["false", "true"]
    check_dict["log_dyn_requests"] = ["false", "true"]
    check_dict["print_dyn"] = ["false", "true"]
    check_dict["log_privilege_changes"] = ["false", "true"]
    check_dict["log_changes_only"] = ["false", "true"]
    check_dict["log_services"] = ["false", "false"]
    check_dict["reset_counters"] = ["true", "true"]
    check_dict["print_hostname"] = ["true", "true"]
    check_dict["cancel_on_error"] = ["true", "true"]
    check_dict["log_message"] = ["true", "true"]
    check_dict["print_security_type"] = ["true", "true"]
    check_dict["explain_plan"] = ["true", "true"]
    check_dict["print_security_level"] = ["true", "true"]
    check_dict["log_sweep"] = ["true", "true"]
    check_dict["include_user_filter"] = ["ship", "0"]
    check_dict["exclude_user_filter"] = ["819", "0"]
    check_dict["include_process_filter"] = ["14", "0"]
    check_dict["exclude_process_filter"] = ["true", "0"]
    check_dict["include_filter"] = ["0", "ship"]
    check_dict["exclude_filter"] = ["0", "819"]
    check_dict["connection_id"] = ["0", "14"]
    check_dict["log_filename"] = ["0", "true"]
    check_dict["max_log_size"] = ["1024", "0"]
    check_dict["time_threshold"] = ["100", "100"]
    check_dict["max_sql_length"] = ["4096", "0"]
    check_dict["max_blr_length"] = ["8192", "0"]
    check_dict["max_dyn_length"] = ["0", "1024"]
    check_dict["max_arg_length"] = ["0", "2048"]
    check_dict["max_arg_count"] = ["0", "4096"]
    check_dict["exclude_gds_codes"] = ["0", "512"]
    check_dict["include_gds_codes"] = ["512", "0"]
    check_dict["archive_directory"] = ["0", "path"]

    with open(conf_path, "r") as f:
        for line in f:
            splited_current_line = list(line.split())
            if len(splited_current_line) == 3:
                first, equal, second = splited_current_line
                assert check_dict[first][number] == second

def create_objects():
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
        with fdb.connect("localhost:employee.fdb", user="SYSDBA", password="masterkey") as con:
            #add objects
            con.execute_immediate("CREATE GLOBAL TEMPORARY TABLE NEW_GLOBAL_TEMPORARY_1 (TEST INTEGER) ON COMMIT DELETE ROWS;")
            con.execute_immediate("""
CREATE OR ALTER TRIGGER NEW_DB_TRIGGER
ACTIVE ON CONNECT POSITION 0
AS
BEGIN
END
""")        
            con.execute_immediate("""
DECLARE EXTERNAL FUNCTION NEW_UDF
RETURNS
BIGINT
ENTRY_POINT '123' MODULE_NAME '123'
""")
            con.execute_immediate("CREATE ROLE TEST_ROLE;")       
            con.execute_immediate("create collation iso8859_1_unicode for iso8859_1")
            #add comments
            con.execute_immediate("COMMENT ON DOMAIN EMPNO IS 'comment'")
            con.execute_immediate("COMMENT ON TABLE EMPLOYEE IS 'comment'")
            con.execute_immediate("COMMENT ON TABLE NEW_GLOBAL_TEMPORARY_1 IS 'comment'")
            con.execute_immediate("COMMENT ON VIEW PHONE_LIST IS 'comment'")
            con.execute_immediate("COMMENT ON PROCEDURE ADD_EMP_PROJ IS 'comment'")
            con.execute_immediate("COMMENT ON TRIGGER POST_NEW_ORDER IS 'comment'")
            con.execute_immediate("COMMENT ON TRIGGER NEW_DB_TRIGGER IS 'comment'")
            con.execute_immediate("COMMENT ON SEQUENCE CUST_NO_GEN IS 'comment';")
            con.execute_immediate("COMMENT ON EXCEPTION CUSTOMER_CHECK IS 'comment'")
            con.execute_immediate("COMMENT ON EXTERNAL FUNCTION NEW_UDF IS 'comment'")
            con.execute_immediate("COMMENT ON ROLE TEST_ROLE IS 'comment'")
            con.execute_immediate("COMMENT ON INDEX CHANGEX IS 'comment'")
            con.commit()
    else:
        with fdb.connect("localhost:employee.fdb", user="SYSDBA", password="masterkey") as con:
            #add objects
            con.execute_immediate("CREATE GLOBAL TEMPORARY TABLE NEW_GLOBAL_TEMPORARY_1 (TEST INTEGER) ON COMMIT DELETE ROWS;")
            con.execute_immediate("""
CREATE OR ALTER FUNCTION NEW_FUNC
RETURNS VARCHAR(5)
AS
begin
RETURN 'five';
end
""")
            con.execute_immediate("CREATE PACKAGE NEW_PACK AS BEGIN END;")
            con.execute_immediate("RECREATE PACKAGE BODY NEW_PACK AS BEGIN END;")
            con.execute_immediate("""
CREATE OR ALTER TRIGGER NEW_DDL_TRIGGER
ACTIVE BEFORE ANY DDL STATEMENT  POSITION 0
AS
BEGIN
END
""")
            con.execute_immediate("""
CREATE OR ALTER TRIGGER NEW_DB_TRIGGER
ACTIVE ON CONNECT POSITION 0
AS
BEGIN
END
""")
            con.execute_immediate("""
DECLARE EXTERNAL FUNCTION NEW_UDF
RETURNS
BIGINT
ENTRY_POINT '123' MODULE_NAME '123'
""")
            con.execute_immediate("CREATE ROLE TEST_ROLE;")       
            con.execute_immediate("create collation iso8859_1_unicode for iso8859_1")

            #add comments
            con.execute_immediate("COMMENT ON DOMAIN EMPNO IS 'comment'")
            con.execute_immediate("COMMENT ON TABLE EMPLOYEE IS 'comment'")
            con.execute_immediate("COMMENT ON TABLE NEW_GLOBAL_TEMPORARY_1 IS 'comment'")
            con.execute_immediate("COMMENT ON VIEW PHONE_LIST IS 'comment'")
            con.execute_immediate("COMMENT ON PROCEDURE ADD_EMP_PROJ IS 'comment'")
            con.execute_immediate("COMMENT ON FUNCTION NEW_FUNC IS 'comment'")
            con.execute_immediate("COMMENT ON PACKAGE NEW_PACK IS 'comment'")
            con.execute_immediate("COMMENT ON TRIGGER POST_NEW_ORDER IS 'comment'")
            con.execute_immediate("COMMENT ON TRIGGER NEW_DDL_TRIGGER IS 'comment'")
            con.execute_immediate("COMMENT ON TRIGGER NEW_DB_TRIGGER IS 'comment'")
            con.execute_immediate("COMMENT ON SEQUENCE CUST_NO_GEN IS 'comment';")
            con.execute_immediate("COMMENT ON EXCEPTION CUSTOMER_CHECK IS 'comment'")
            con.execute_immediate("COMMENT ON EXTERNAL FUNCTION NEW_UDF IS 'comment'")
            con.execute_immediate("COMMENT ON ROLE TEST_ROLE IS 'comment'")
            con.execute_immediate("COMMENT ON INDEX CHANGEX IS 'comment'")
            con.commit()

            # if rdb5:
            # con.execute_immediate("CREATE TABLESPACE NEW_TABLESPACE_1 FILE 'file.ts';")
            #             con.execute_immediate("""                                 
            # CREATE JOB NEW_JOB_EXPORT
            # '13 17 * * *'
            # INACTIVE
            # START DATE NULL
            # END DATE NULL
            # AS
            # begin
            # end
            # """)
        

def delete_objects(rdb5: bool):
    with fdb.connect("employee") as con:
        if rdb5:
            # con.execute_immediate("DROP TABLESPACE NEW_TABLESPACE_1;")
            con.execute_immediate("DROP JOB NEW_JOB_EXPORT;")
            con.commit()

def delete_files(files: list):
	for file in files:
		if os.path.exists(file):
			os.remove(file)

def create_database(script_path: str, base_path: str):
    home_directory, version, srv_version = get_server_info()
    with open(script_path, "r") as file:
        context = file.read()
    
    context = context.replace("employee.fdb", base_path)
    
    with open(script_path, "w") as file:
        file.write(context)
    
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
    con = fdb.create_database(database=base_path, user='SYSDBA', password='masterkey')
    con.close()
    bin_dir = "bin/" if platform.system() == "Linux" or is_rdb26() else ""
    subprocess.call([f"{home_directory}{bin_dir}isql{get_exe()}",  "-q", "-i", f"\"{script_path}\""])

def build_procedure():
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
    with fdb.connect("localhost:employee.fdb", user="SYSDBA", password="masterkey") as con:
        with con.cursor() as cur:
            results = cur.execute("SELECT RDB$KEYWORD_NAME FROM RDB$KEYWORDS WHERE RDB$KEYWORD_RESERVED='false'")
            res = results.fetchall()
        
        create_procudere_script = "CREATE OR ALTER PROCEDURE DECLARE_KEYWORDS ( NUM_ENTRIES INTEGER ) AS"
        for row in res:
            create_procudere_script += f" DECLARE {row[0]} INTEGER;\n"
        create_procudere_script += " BEGIN END;"

    execute_immediate(create_procudere_script)
    return create_procudere_script


def check_xlsx(xlsx_path: str):
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = openpyxl.load_workbook(xlsx_path)
    
    worksheet = workbook.active
    result = ""
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            result += f"{col[i].value}\t" 
        result += "\n"
    
    return result

def add_rows():
    execute_immediate("CREATE TABLE TEST_TABLE (ID BIGINT)")
    import firebird.driver as fdb
    if is_rdb26():
        import fdb
        load_api()
    with fdb.connect("localhost:employee.fdb", user="SYSDBA", password="masterkey") as con:
        for i in range(1048576):
            con.execute_immediate(f"INSERT INTO TEST_TABLE VALUES ({i})")
    
        con.commit()

def is_rdb26():
    DBMS = os.environ.get('DBMS')
    return True if DBMS == "rdb26" else False

def load_api():
    import fdb
    home_directory, _, _ = get_server_info()
    fd_lib = "lib/libfbclient.so" if platform.system() == "Linux" else "bin/fbclient.dll"
    fdb.load_api(home_directory + fd_lib)

def get_exe():
    return "" if platform.system() == "Linux" else ".exe"

def get_user_for_ssh():
    user = "reduser" if platform.system() == "Linux" else "jenkins"
    password = "1" if platform.system() == "Linux" else "jenkins"
    return  user, password